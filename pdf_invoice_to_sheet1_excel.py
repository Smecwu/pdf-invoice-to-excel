from __future__ import annotations

import argparse
import re
from copy import copy
from dataclasses import dataclass
from datetime import date
from decimal import Decimal
from pathlib import Path

import pdfplumber
from openpyxl import load_workbook


HEADER_ROW = 2
DATA_START_ROW = 3

INVOICE_NO = "\u53d1\u7968\u53f7\u7801"
ISSUE_DATE = "\u5f00\u7968\u65e5\u671f"
TOTAL_WITH_TAX = "\u4ef7\u7a0e\u5408\u8ba1"
LOWERCASE_AMOUNT = "\uff08\u5c0f\u5199\uff09"
REMARK_FIRST = "\u5907"
REMARK_SECOND = "\u6ce8"
MONTH_INVOICE_TITLE = "\u5e74{month}\u6708\u4efd\u53d1\u7968"


@dataclass(frozen=True)
class InvoiceItem:
    material: str
    quantity: Decimal
    amount: Decimal
    tax: Decimal


@dataclass(frozen=True)
class Invoice:
    source_pdf: Path
    invoice_no: str
    invoice_date: date
    total: Decimal
    remark: str
    items: list[InvoiceItem]


def parse_decimal(value: str) -> Decimal:
    return Decimal(value.replace(",", "").strip())


def decimal_to_excel(value: Decimal) -> int | float:
    if value == value.to_integral_value():
        return int(value)
    return float(value)


def extract_text(pdf_path: Path) -> str:
    with pdfplumber.open(pdf_path) as pdf:
        return "\n".join(page.extract_text() or "" for page in pdf.pages)


def parse_invoice_date(text: str) -> date:
    match = re.search(
        rf"{ISSUE_DATE}[:\uff1a]\s*(\d{{4}})\u5e74(\d{{1,2}})\u6708(\d{{1,2}})\u65e5",
        text,
    )
    if not match:
        raise ValueError("invoice date not found")
    year, month, day = (int(part) for part in match.groups())
    return date(year, month, day)


def parse_remark(text: str) -> str:
    match = re.search(
        rf"{TOTAL_WITH_TAX}.*?{LOWERCASE_AMOUNT}\s*[\u00a5\uffe5]?\s*[\d,.]+\s*"
        rf"\n(?P<remark>.*?)\n{REMARK_FIRST}\s*\n?{REMARK_SECOND}",
        text,
        flags=re.S,
    )
    if not match:
        return ""
    lines = [line.strip() for line in match.group("remark").splitlines() if line.strip()]
    return ",".join(lines)


def parse_material(prefix: str) -> str:
    candidates = re.findall(r"[A-Z][A-Z0-9]{1,}(?:-[A-Z0-9]+)*", prefix)
    if not candidates:
        raise ValueError(f"material code not found: {prefix}")
    return candidates[-1]


def parse_items(text: str) -> list[InvoiceItem]:
    item_line = re.compile(
        r"^(?P<prefix>\*.+?)\s+"
        r"(?P<unit>[A-Za-z]{1,5}|[\u4e00-\u9fff]{1,4})\s+"
        r"(?P<quantity>-?\d+(?:\.\d+)?)\s+"
        r"(?P<unit_price>-?\d+(?:\.\d+)?)\s+"
        r"(?P<amount>-?\d+(?:\.\d+)?)\s+"
        r"(?P<rate>\d+(?:\.\d+)?%|\u514d\u7a0e|\u4e0d\u5f81\u7a0e)\s+"
        r"(?P<tax>-?\d+(?:\.\d+)?)$"
    )
    items: list[InvoiceItem] = []

    for raw_line in text.splitlines():
        line = re.sub(r"\s+", " ", raw_line.strip())
        if not line.startswith("*"):
            continue
        match = item_line.match(line)
        if not match:
            continue
        items.append(
            InvoiceItem(
                material=parse_material(match.group("prefix")),
                quantity=parse_decimal(match.group("quantity")),
                amount=parse_decimal(match.group("amount")),
                tax=parse_decimal(match.group("tax")),
            )
        )

    if not items:
        raise ValueError("invoice item rows not found")
    return items


def parse_invoice(pdf_path: Path) -> Invoice:
    text = extract_text(pdf_path)

    invoice_no_match = re.search(rf"{INVOICE_NO}[:\uff1a]\s*(\d+)", text)
    if not invoice_no_match:
        raise ValueError("invoice number not found")

    total_match = re.search(
        rf"{TOTAL_WITH_TAX}.*?{LOWERCASE_AMOUNT}\s*[\u00a5\uffe5]?\s*([\d,.]+)",
        text,
        flags=re.S,
    )
    if not total_match:
        raise ValueError("invoice total not found")

    return Invoice(
        source_pdf=pdf_path,
        invoice_no=invoice_no_match.group(1),
        invoice_date=parse_invoice_date(text),
        total=parse_decimal(total_match.group(1)),
        remark=parse_remark(text),
        items=parse_items(text),
    )


def find_invoice_pdfs(month_dir: Path, include_merged: bool) -> list[Path]:
    pdfs = sorted(month_dir.glob("*.pdf"), key=lambda path: path.name.lower())
    if include_merged:
        return pdfs
    return [path for path in pdfs if "merge" not in path.stem.lower()]


def unique_by_invoice_no(invoices: list[Invoice]) -> tuple[list[Invoice], list[Invoice]]:
    unique: list[Invoice] = []
    skipped: list[Invoice] = []
    seen: set[str] = set()

    for invoice in invoices:
        if invoice.invoice_no in seen:
            skipped.append(invoice)
            continue
        seen.add(invoice.invoice_no)
        unique.append(invoice)

    return unique, skipped


def row_style_snapshot(ws, row_number: int) -> tuple[list[dict[str, object]], float | None]:
    styles: list[dict[str, object]] = []
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row_number, col)
        styles.append(
            {
                "style": copy(cell._style),
                "number_format": cell.number_format,
                "alignment": copy(cell.alignment),
                "font": copy(cell.font),
                "fill": copy(cell.fill),
                "border": copy(cell.border),
            }
        )
    return styles, ws.row_dimensions[row_number].height


def apply_row_style(ws, row_number: int, styles: list[dict[str, object]], height: float | None) -> None:
    for col, style in enumerate(styles, start=1):
        cell = ws.cell(row_number, col)
        cell._style = copy(style["style"])
        cell.number_format = style["number_format"]
        cell.alignment = copy(style["alignment"])
        cell.font = copy(style["font"])
        cell.fill = copy(style["fill"])
        cell.border = copy(style["border"])
    ws.row_dimensions[row_number].height = height


def clear_data_rows(ws) -> None:
    if ws.max_row >= DATA_START_ROW:
        ws.delete_rows(DATA_START_ROW, ws.max_row - DATA_START_ROW + 1)


def infer_title(month_dir: Path) -> str:
    match = re.search(r"(\d{4})[.\-_\u5e74]?(\d{1,2})", month_dir.name)
    if not match:
        return ""
    year, month = match.groups()
    return f"{year}{MONTH_INVOICE_TITLE.format(month=int(month))}"


def write_sheet1_only(
    template_path: Path,
    output_path: Path,
    invoices: list[Invoice],
    title: str | None,
) -> int:
    workbook = load_workbook(template_path)
    ws = workbook[workbook.sheetnames[0]]

    for extra_ws in list(workbook.worksheets[1:]):
        workbook.remove(extra_ws)
    ws.title = "Sheet1"

    sample_styles, sample_height = row_style_snapshot(ws, DATA_START_ROW)
    clear_data_rows(ws)

    if title:
        ws.cell(1, 1).value = title

    row = DATA_START_ROW
    for invoice in invoices:
        for item_index, item in enumerate(invoice.items):
            apply_row_style(ws, row, sample_styles, sample_height)
            invoice_no_cell = ws.cell(row, 1)
            invoice_no_cell.value = f'=MID("x{invoice.invoice_no}",2,99)'
            invoice_no_cell.number_format = "@"
            ws.cell(row, 2).value = invoice.invoice_date
            ws.cell(row, 3).value = item.material
            ws.cell(row, 3).number_format = "@"
            ws.cell(row, 4).value = decimal_to_excel(item.quantity)
            ws.cell(row, 5).value = decimal_to_excel(item.amount)
            ws.cell(row, 6).value = decimal_to_excel(item.tax)
            ws.cell(row, 7).value = (
                decimal_to_excel(invoice.total) if item_index == len(invoice.items) - 1 else None
            )
            ws.cell(row, 8).value = invoice.remark
            ws.cell(row, 8).number_format = "@"
            row += 1

    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width or 0, 24)
    ws.column_dimensions["C"].width = max(ws.column_dimensions["C"].width or 0, 22)
    ws.column_dimensions["H"].width = max(ws.column_dimensions["H"].width or 0, 28)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return row - DATA_START_ROW


def resolve_path(path_text: str, base_dir: Path) -> Path:
    path = Path(path_text)
    if path.is_absolute():
        return path
    return (base_dir / path).resolve()


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Extract Chinese e-invoice PDFs into one Sheet1 Excel workbook."
    )
    parser.add_argument("--month-dir", required=True, help="Folder containing invoice PDFs.")
    parser.add_argument("--template", required=True, help="Template xlsx path. Only its first sheet is used.")
    parser.add_argument("--output", required=True, help="Output xlsx path.")
    parser.add_argument("--title", default=None, help="Optional title for cell A1.")
    parser.add_argument("--include-merged", action="store_true", help="Also read PDF files with 'merge' in filename.")
    parser.add_argument("--keep-duplicates", action="store_true", help="Keep duplicate invoice numbers.")
    parser.add_argument("--dry-run", action="store_true", help="Parse only; do not write Excel.")
    return parser


def main() -> None:
    args = build_parser().parse_args()
    root = Path(__file__).resolve().parent
    month_dir = resolve_path(args.month_dir, root)
    template_path = resolve_path(args.template, month_dir)
    output_path = resolve_path(args.output, month_dir)

    invoices: list[Invoice] = []
    errors: list[tuple[Path, str]] = []
    for pdf_path in find_invoice_pdfs(month_dir, args.include_merged):
        try:
            invoices.append(parse_invoice(pdf_path))
        except Exception as exc:
            errors.append((pdf_path, str(exc)))

    invoices.sort(key=lambda item: (item.invoice_date, item.invoice_no, item.source_pdf.name))
    skipped_duplicates: list[Invoice] = []
    if not args.keep_duplicates:
        invoices, skipped_duplicates = unique_by_invoice_no(invoices)

    item_count = sum(len(invoice.items) for invoice in invoices)
    total_amount = sum((invoice.total for invoice in invoices), Decimal("0"))
    print(f"parsed invoices: {len(invoices)}")
    print(f"parsed item rows: {item_count}")
    print(f"sum total with tax: {total_amount:.2f}")

    if skipped_duplicates:
        print(f"skipped duplicate PDFs: {len(skipped_duplicates)}")
        for invoice in skipped_duplicates[:20]:
            print(f"- {invoice.source_pdf.name}: {invoice.invoice_no}")

    if errors:
        print(f"parse errors: {len(errors)}")
        for pdf_path, message in errors[:20]:
            print(f"- {pdf_path.name}: {message}")

    if args.dry_run:
        return
    if errors:
        raise SystemExit("Stopped because one or more PDFs could not be parsed.")

    title = args.title if args.title is not None else infer_title(month_dir)
    written_rows = write_sheet1_only(template_path, output_path, invoices, title)
    print(f"saved: {output_path}")
    print(f"written rows: {written_rows}")


if __name__ == "__main__":
    main()
